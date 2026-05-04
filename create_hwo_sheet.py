"""
create_hwo_sheet.py  (Module 2 — Tidal Locking)
=================================================
Run this once to generate hwo_targets.xlsx from the published HWO
dynamical viability target list (Table 1, Quirrenbach et al. style).

    python module_2_tidal_rotation/create_hwo_sheet.py

Columns written
---------------
planet_name, star_name, star_common, M_star_Msun, Teff_K,
M_planet_MJ, M_planet_Mearth, a_AU, eccentricity,
system_age_gyr, planet_type, has_rv_data, notes

planet_type is inferred from M_planet_Mearth:
  < 10   → rocky
  10–100 → ice_giant
  > 100  → gas_giant

P_orb_days is NOT stored — batch_processor.py derives it from Kepler's law
so it stays consistent with whatever a_AU and M_star are used.
"""

import os
import numpy as np

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

MJ_TO_ME = 317.83   # Jupiter mass → Earth mass

# ── Planet data from Table 1 ──────────────────────────────────────────────────
# Format: (star_id, star_common, M_star_Msun, Teff_K, system_age_gyr,
#           planet_letter, M_planet_MJ, a_AU, eccentricity, notes)
# System ages from literature (see references in batch_processor.py)

RAW = [
    # HD 3651
    ("HD 3651",   "54 Piscium",   0.799, 5221, 6.8,  "b",  0.228,   0.295,    0.645, ""),
    # HD 9826
    ("HD 9826",   "ups And",      1.29,  6156, 3.3,  "b",  0.675,   0.05914,  0.0069,""),
    ("HD 9826",   "ups And",      1.29,  6156, 3.3,  "c",  1.965,   0.8265,   0.266, ""),
    ("HD 9826",   "ups And",      1.29,  6156, 3.3,  "d",  4.1,     2.517,    0.294, ""),
    # HD 10647
    ("HD 10647",  "q1 Eridani",   1.11,  6218, 1.5,  "b",  0.94,    2.015,    0.15,  ""),
    # HD 10700 (tau Ceti)
    ("HD 10700",  "tau Ceti",     0.78,  5333, 5.8,  "g",  0.0055,  0.133,    0.06,  "min mass"),
    ("HD 10700",  "tau Ceti",     0.78,  5333, 5.8,  "h",  0.0058,  0.243,    0.23,  "min mass"),
    ("HD 10700",  "tau Ceti",     0.78,  5333, 5.8,  "e",  0.0124,  0.538,    0.18,  "min mass"),
    ("HD 10700",  "tau Ceti",     0.78,  5333, 5.8,  "f",  0.0124,  1.334,    0.16,  "min mass"),
    # HD 17051
    ("HD 17051",  "iota Hor",     1.34,  6167, 0.6,  "b",  2.27,    0.92,     0.14,  ""),
    # HD 20794 (82 Eridani)
    ("HD 20794",  "82 Eridani",   0.813, 5401, 7.1,  "b",  0.0089,  0.127,    0.27,  "min mass"),
    ("HD 20794",  "82 Eridani",   0.813, 5401, 7.1,  "c",  0.0079,  0.225,    0.17,  "min mass"),
    ("HD 20794",  "82 Eridani",   0.813, 5401, 7.1,  "d",  0.0111,  0.364,    0.25,  "min mass"),
    ("HD 20794",  "82 Eridani",   0.813, 5401, 7.1,  "e",  0.015,   0.509,    0.29,  "min mass"),
    # HD 22049 (eps Eri)
    ("HD 22049",  "eps Eridani",  0.81,  5020, 0.7,  "b",  0.651,   3.5,      0.044, ""),
    # HD 26965 (40 Eri)
    ("HD 26965",  "40 Eridani",   0.78,  5072, 5.6,  "b",  0.0267,  0.219,    0.04,  "min mass"),
    # HD 33564
    ("HD 33564",  "kappa Cam",    1.25,  6250, 2.0,  "b",  9.1,     1.1,      0.34,  ""),
    # HD 39091 (pi Men)
    ("HD 39091",  "pi Mensae",    1.07,  5998, 2.6,  "c",  0.0114,  0.06805,  0.0,   "true mass"),
    ("HD 39091",  "pi Mensae",    1.07,  5998, 2.6,  "d",  0.0421,  0.499,    0.22,  "min mass"),
    ("HD 39091",  "pi Mensae",    1.07,  5998, 2.6,  "b",  12.6,    3.2826,   0.6396,""),
    # HD 69830
    ("HD 69830",  "HD 69830",     0.86,  5385, 7.9,  "b",  0.0321,  0.0785,   0.1,   "min mass"),
    ("HD 69830",  "HD 69830",     0.86,  5385, 7.9,  "c",  0.0371,  0.186,    0.13,  "min mass"),
    ("HD 69830",  "HD 69830",     0.86,  5385, 7.9,  "d",  0.057,   0.63,     0.07,  "min mass"),
    # HD 75732 (55 Cancri)
    ("HD 75732",  "55 Cancri",    0.905, 5172, 10.2, "e",  0.0251,  0.01544,  0.05,  "true mass; lava world"),
    ("HD 75732",  "55 Cancri",    0.905, 5172, 10.2, "b",  0.8036,  0.1134,   0.0,   ""),
    ("HD 75732",  "55 Cancri",    0.905, 5172, 10.2, "c",  0.1611,  0.2373,   0.03,  ""),
    ("HD 75732",  "55 Cancri",    0.905, 5172, 10.2, "f",  0.1503,  0.7708,   0.08,  ""),
    ("HD 75732",  "55 Cancri",    0.905, 5172, 10.2, "d",  3.12,    5.957,    0.13,  ""),
    # HD 95128 (47 UMa)
    ("HD 95128",  "47 UMa",       1.06,  5872, 6.0,  "b",  2.53,    2.1,      0.032, ""),
    ("HD 95128",  "47 UMa",       1.06,  5872, 6.0,  "c",  0.54,    3.6,      0.098, ""),
    ("HD 95128",  "47 UMa",       1.06,  5872, 6.0,  "d",  1.64,    11.6,     0.16,  ""),
    # HD 95735
    ("HD 95735",  "HD 95735",     0.3899,3712, 0.5,  "b",  0.0085,  0.07879,  0.063, "M dwarf; min mass"),
    ("HD 95735",  "HD 95735",     0.3899,3712, 0.5,  "c",  0.0428,  2.94,     0.132, "M dwarf; min mass"),
    # HD 102365
    ("HD 102365", "HD 102365",    0.85,  5630, 5.6,  "b",  0.0503,  0.46,     0.34,  "min mass"),
    # HD 114613
    ("HD 114613", "HD 114613",    1.27,  5641, 4.9,  "b",  0.357,   5.34,     0.458, ""),
    # HD 115404A
    ("HD 115404A","HD 115404A",   0.83,  5019, 0.8,  "b",  0.097,   0.088,    0.232, "min mass"),
    ("HD 115404A","HD 115404A",   0.83,  5019, 0.8,  "c",  10.319,  11.364,   0.211, ""),
    # HD 115617 (61 Vir)
    ("HD 115617", "61 Vir",       0.942, 5577, 6.4,  "b",  0.016,   0.050201, 0.12,  "min mass"),
    ("HD 115617", "61 Vir",       0.942, 5577, 6.4,  "c",  0.057,   0.2175,   0.14,  "min mass"),
    ("HD 115617", "61 Vir",       0.942, 5577, 6.4,  "d",  0.072,   0.476,    0.35,  "min mass"),
    # HD 136352 (nu2 Lupi)
    ("HD 136352", "nu2 Lupi",     0.87,  5664, 4.0,  "b",  0.0149,  0.0964,   0.0,   "true mass; transiting"),
    ("HD 136352", "nu2 Lupi",     0.87,  5664, 4.0,  "c",  0.0354,  0.1721,   0.0,   "true mass; transiting"),
    ("HD 136352", "nu2 Lupi",     0.87,  5664, 4.0,  "d",  0.0278,  0.425,    0.0,   "true mass; transiting"),
    # HD 141004 (lambda Ser)
    ("HD 141004", "lambda Ser",   1.05,  5885, 7.0,  "b",  0.0428,  0.1238,   0.16,  "min mass"),
    # HD 140901
    ("HD 140901", "HD 140901",    0.99,  5586, 4.0,  "b",  0.0503,  0.085,    0.472, "min mass"),
    ("HD 140901", "HD 140901",    0.99,  5586, 4.0,  "c",  6.284,   7.421,    0.607, ""),
    # HD 143761 (rho CrB)
    ("HD 143761", "rho CrB",      0.95,  5817, 11.5, "e",  0.0119,  0.1061,   0.126, "min mass"),
    ("HD 143761", "rho CrB",      0.95,  5817, 11.5, "b",  1.093,   0.2245,   0.038, ""),
    ("HD 143761", "rho CrB",      0.95,  5817, 11.5, "c",  0.0887,  0.4206,   0.096, "min mass"),
    ("HD 143761", "rho CrB",      0.95,  5817, 11.5, "d",  0.068,   0.827,    0.0,   "min mass"),
    # HD 147513
    ("HD 147513", "62 G. Sco",    1.11,  5883, 0.6,  "b",  1.21,    1.32,     0.26,  ""),
    # HD 160691 (mu Ara)
    ("HD 160691", "mu Ara",       1.13,  5773, 6.3,  "d",  0.033,   0.0923,   0.16,  "min mass"),
    ("HD 160691", "mu Ara",       1.13,  5773, 6.3,  "e",  0.439,   0.9296,   0.091, ""),
    ("HD 160691", "mu Ara",       1.13,  5773, 6.3,  "b",  1.665,   1.5224,   0.036, ""),
    ("HD 160691", "mu Ara",       1.13,  5773, 6.3,  "c",  1.873,   5.0937,   0.022, ""),
    # HD 189567
    ("HD 189567", "HD 189567",    0.83,  5726, 7.0,  "b",  0.0267,  0.111,    0.0,   "min mass"),
    ("HD 189567", "HD 189567",    0.83,  5726, 7.0,  "c",  0.022,   0.197,    0.16,  "min mass"),
    # HD 190360
    ("HD 190360", "HD 190360",    0.99,  5537, 7.9,  "c",  0.0677,  0.1294,   0.165, "min mass"),
    ("HD 190360", "HD 190360",    0.99,  5537, 7.9,  "b",  1.492,   3.955,    0.3274,""),
    # HD 192310
    ("HD 192310", "HD 192310",    0.8,   5166, 7.7,  "b",  0.0532,  0.32,     0.13,  "min mass"),
    ("HD 192310", "HD 192310",    0.8,   5166, 7.7,  "c",  0.076,   1.18,     0.32,  "min mass"),
    # HD 209100 (eps Indi A)
    ("HD 209100", "eps Indi A",   0.754, 4611, 3.7,  "b",  3.25,    11.55,    0.26,  ""),
    # HD 217987
    ("HD 217987", "HD 217987",    0.489, 3688, 4.5,  "b",  0.0132,  0.068,    0.0,   "M dwarf; min mass"),
    ("HD 217987", "HD 217987",    0.489, 3688, 4.5,  "c",  0.0239,  0.12,     0.0,   "M dwarf; min mass"),
    # HD 219134
    ("HD 219134", "HD 219134",    0.794, 4913, 10.8, "b",  0.012,   0.038474, 0.0,   "transiting; min mass"),
    ("HD 219134", "HD 219134",    0.794, 4913, 10.8, "c",  0.011,   0.064816, 0.0,   "transiting; min mass"),
    ("HD 219134", "HD 219134",    0.794, 4913, 10.8, "f",  0.028,   0.14574,  0.0,   "min mass"),
    ("HD 219134", "HD 219134",    0.794, 4913, 10.8, "d",  0.067,   0.23508,  0.0,   "min mass"),
    ("HD 219134", "HD 219134",    0.794, 4913, 10.8, "g",  0.0346,  0.3753,   0.0,   "min mass"),
    ("HD 219134", "HD 219134",    0.794, 4913, 10.8, "h",  0.34,    3.11,     0.06,  ""),
]


def _planet_type(M_earth):
    if M_earth < 10.0:
        return "rocky"
    elif M_earth < 100.0:
        return "ice_giant"
    return "gas_giant"


def build_rows():
    rows = []
    for (star, common, Mstar, Teff, age, letter, MJ, a, ecc, notes) in RAW:
        ME     = MJ * MJ_TO_ME
        ptype  = _planet_type(ME)
        pname  = f"{star} {letter}"
        rows.append({
            "planet_name":       pname,
            "star_name":         star,
            "star_common":       common,
            "M_star_Msun":       Mstar,
            "Teff_K":            Teff,
            "M_planet_MJ":       round(MJ, 6),
            "M_planet_Mearth":   round(ME, 4),
            "a_AU":              a,
            "eccentricity":      ecc,
            "system_age_gyr":    age,
            "planet_type":       ptype,
            "has_rv_data":       False,
            "notes":             notes,
        })
    return rows


def write_xlsx(rows, out_path):
    if not HAS_OPENPYXL:
        print("openpyxl not installed. Writing CSV instead.")
        import csv
        csv_path = out_path.replace(".xlsx", ".csv")
        with open(csv_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        print(f"  Saved: {csv_path}")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HWO Targets"

    headers = list(rows[0].keys())

    # ── Header style ────────────────────────────────────────────────
    hdr_fill   = PatternFill("solid", fgColor="1A3A5C")
    hdr_font   = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side  = Side(style="thin", color="AAAAAA")
    thin_border= Border(left=thin_side, right=thin_side,
                        top=thin_side, bottom=thin_side)

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font   = hdr_font
        cell.fill   = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    # ── Data rows ───────────────────────────────────────────────────
    rocky_fill = PatternFill("solid", fgColor="E8F5E9")   # light green
    ice_fill   = PatternFill("solid", fgColor="E3F2FD")   # light blue
    gas_fill   = PatternFill("solid", fgColor="FFF3E0")   # light orange

    fill_map = {"rocky": rocky_fill, "ice_giant": ice_fill, "gas_giant": gas_fill}

    for row_idx, row in enumerate(rows, start=2):
        ptype = row.get("planet_type", "rocky")
        row_fill = fill_map.get(ptype, rocky_fill)
        for col_idx, key in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[key])
            cell.fill   = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if isinstance(row[key], float):
                cell.number_format = "0.000000"

    # ── Column widths ───────────────────────────────────────────────
    col_widths = {
        "planet_name": 18, "star_name": 14, "star_common": 16,
        "M_star_Msun": 14, "Teff_K": 10, "M_planet_MJ": 14,
        "M_planet_Mearth": 18, "a_AU": 12, "eccentricity": 14,
        "system_age_gyr": 17, "planet_type": 14,
        "has_rv_data": 13, "notes": 30,
    }
    for col_idx, key in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(key, 14)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Legend sheet ────────────────────────────────────────────────
    ls = wb.create_sheet("Column Guide")
    guide = [
        ("Column", "Description", "Units / Values"),
        ("planet_name",      "Unique planet identifier",                  "e.g. HD 219134 b"),
        ("star_name",        "HD catalogue name of host star",            "string"),
        ("star_common",      "Common name",                               "string"),
        ("M_star_Msun",      "Stellar mass",                              "Solar masses"),
        ("Teff_K",           "Stellar effective temperature",             "Kelvin"),
        ("M_planet_MJ",      "Planet minimum mass (M sin i for most)",    "Jupiter masses"),
        ("M_planet_Mearth",  "Same, converted",                           "Earth masses"),
        ("a_AU",             "Semi-major axis",                           "AU"),
        ("eccentricity",     "Orbital eccentricity",                      "0 – 1"),
        ("system_age_gyr",   "Estimated system age (literature)",         "Gyr"),
        ("planet_type",      "Used to select Q and k2 defaults",          "rocky / ice_giant / gas_giant"),
        ("has_rv_data",      "Set True to trigger RM MCMC for this row",  "True / False"),
        ("notes",            "Any remarks",                               "string"),
    ]
    for r, row in enumerate(guide, start=1):
        for c, val in enumerate(row, start=1):
            cell = ls.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True)

    wb.save(out_path)
    print(f"  Saved: {out_path}  ({len(rows)} planets)")


if __name__ == "__main__":
    here    = os.path.dirname(os.path.abspath(__file__))
    out     = os.path.join(here, "hwo_targets.xlsx")
    rows    = build_rows()
    write_xlsx(rows, out)
    print(f"  Rocky planets: {sum(1 for r in rows if r['planet_type']=='rocky')}")
    print(f"  Ice giants:    {sum(1 for r in rows if r['planet_type']=='ice_giant')}")
    print(f"  Gas giants:    {sum(1 for r in rows if r['planet_type']=='gas_giant')}")
